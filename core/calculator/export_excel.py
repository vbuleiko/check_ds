"""
Модуль экспорта объёмов работ в Excel.

Генерирует файл по образцу Приложения №12 с помесячными объёмами.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Optional
import calendar

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from db.models import Contract, RouteParams
from core.calculator.kilometers import calculate_route_period, calculate_contract_period
from core.calculator.price import get_coefficients_for_date, get_capacities, preload_price_data


# Названия месяцев
MONTH_NAMES = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
]


def normalize_route(route: str) -> str:
    """
    Нормализует название маршрута.
    Убирает точку в конце: 315. -> 315
    """
    return route.rstrip(".")


def get_contract_routes(session: Session, contract_id: int, normalize: bool = True) -> list[str]:
    """
    Получает список всех маршрутов контракта.

    Args:
        session: Сессия БД
        contract_id: ID контракта
        normalize: Если True, нормализует маршруты (315. -> 315)
    """
    route_records = session.query(RouteParams.route).filter(
        RouteParams.contract_id == contract_id
    ).distinct().all()

    routes = [r[0] for r in route_records]

    # Нормализуем маршруты (объединяем 315 и 315.)
    if normalize:
        normalized = set()
        for route in routes:
            normalized.add(normalize_route(route))
        routes = list(normalized)

    # Сортируем маршруты: числовые первые, затем с буквами
    def sort_key(route: str) -> tuple:
        # Извлекаем числовую часть
        num_part = ""
        suffix = ""
        for i, ch in enumerate(route):
            if ch.isdigit():
                num_part += ch
            elif ch == '.':
                continue  # Пропускаем точки при сортировке
            else:
                suffix = route[i:]
                break

        try:
            num = float(num_part) if num_part else 9999
        except ValueError:
            num = 9999

        return (num, suffix)

    return sorted(routes, key=sort_key)


def generate_monthly_volumes_excel(
    session: Session,
    contract_id: int,
    start_year: int = 2026,
    start_month: int = 2,
    end_year: int = 2028,
    end_month: int = 7,
    ds_number: Optional[str] = None,
) -> BytesIO:
    """
    Генерирует Excel файл с помесячными объёмами работ.

    Args:
        session: Сессия БД
        contract_id: ID контракта
        start_year: Начальный год (по умолчанию 2026)
        start_month: Начальный месяц (по умолчанию февраль)
        end_year: Конечный год (по умолчанию 2028)
        end_month: Конечный месяц (по умолчанию июль)
        ds_number: Номер ДС для заголовка

    Returns:
        BytesIO с Excel файлом
    """
    # Получаем контракт
    contract = session.query(Contract).filter(Contract.id == contract_id).first()
    contract_number = contract.number if contract else "?"

    # Получаем нормализованный список маршрутов (для отображения в таблице)
    routes = get_contract_routes(session, contract_id, normalize=True)

    # Получаем оригинальные маршруты (для расчёта)
    original_routes = get_contract_routes(session, contract_id, normalize=False)

    # Маппинг: нормализованный -> список оригинальных
    route_mapping: dict[str, list[str]] = {}
    for orig in original_routes:
        norm = normalize_route(orig)
        if norm not in route_mapping:
            route_mapping[norm] = []
        route_mapping[norm].append(orig)

    # Создаём книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Объёмы работ"

    # Стили
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Одна строка заголовка: № | Год | Месяц | Период | маршрут1 | маршрут2 | ... | Итого
    total_col = len(routes) + 5 + 1  # 4 базовых + маршруты + Итого
    header_row = 1
    for col, header in enumerate(["№", "Год", "Месяц", "Период"], 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for col, route in enumerate(routes, 5):
        cell = ws.cell(row=header_row, column=col, value=route)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    cell = ws.cell(row=header_row, column=total_col, value="Итого")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

    # Генерируем периоды и рассчитываем объёмы
    data_row = header_row + 1
    period_num = 1

    current_year = start_year
    current_month = start_month

    while (current_year < end_year) or (current_year == end_year and current_month <= end_month):
        # Определяем период месяца
        _, last_day = calendar.monthrange(current_year, current_month)
        d_from = date(current_year, current_month, 1)
        d_to = date(current_year, current_month, last_day)

        # Рассчитываем объёмы по всем маршрутам с учётом маппинга нормализованных маршрутов
        # Для каждого нормализованного маршрута суммируем данные всех оригинальных маршрутов
        # Округляем каждый маршрут до 2 знаков перед суммированием
        route_calculations = {}
        for norm_route in routes:
            total_km = 0.0
            for orig_route in route_mapping[norm_route]:
                calc = calculate_route_period(session, contract_id, orig_route, d_from, d_to)
                total_km += round(calc.total_km, 2)
            route_calculations[norm_route] = total_km

        # Заполняем строку
        ws.cell(row=data_row, column=1, value=period_num).border = thin_border
        ws.cell(row=data_row, column=1).alignment = center_align

        ws.cell(row=data_row, column=2, value=current_year).border = thin_border
        ws.cell(row=data_row, column=2).alignment = center_align

        ws.cell(row=data_row, column=3, value=MONTH_NAMES[current_month]).border = thin_border
        ws.cell(row=data_row, column=3).alignment = center_align

        period_str = f"01.{current_month:02d}-{last_day}.{current_month:02d}"
        ws.cell(row=data_row, column=4, value=period_str).border = thin_border
        ws.cell(row=data_row, column=4).alignment = center_align

        # Объёмы по маршрутам (используем суммированные значения)
        row_total = 0.0
        for col, route in enumerate(routes, 5):
            km = round(route_calculations.get(route, 0), 2)
            row_total += km

            cell = ws.cell(row=data_row, column=col, value=km if km > 0 else 0)
            cell.border = thin_border
            cell.alignment = center_align
            cell.number_format = '#,##0.00'

        # Итого по строке
        cell = ws.cell(row=data_row, column=total_col, value=round(row_total, 2))
        cell.border = thin_border
        cell.alignment = center_align
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)

        # Следующий месяц
        period_num += 1
        data_row += 1
        current_month += 1
        if current_month > 12:
            current_month = 1
            current_year += 1

    # Строка "Итого" внизу
    ws.cell(row=data_row, column=1, value="").border = thin_border
    ws.cell(row=data_row, column=2, value="").border = thin_border
    ws.cell(row=data_row, column=3, value="").border = thin_border
    cell = ws.cell(row=data_row, column=4, value="ИТОГО:")
    cell.border = thin_border
    cell.font = Font(bold=True)
    cell.alignment = center_align

    # Суммы по столбцам
    first_data_row = header_row + 1
    last_data_row = data_row - 1

    for col in range(5, total_col + 1):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        cell = ws.cell(row=data_row, column=col, value=formula)
        cell.border = thin_border
        cell.alignment = center_align
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)

    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12

    for col in range(5, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def _build_monthly_sheet(
    ws,
    routes: list[str],
    monthly_rows: list[tuple],
    number_format: str = '#,##0.00',
) -> None:
    """
    Заполняет лист данными помесячных объёмов.

    monthly_rows: list of (period_num, year, month_name, period_str, {route: value})
    """
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    total_col = len(routes) + 5

    # Заголовок
    for col, header in enumerate(["№", "Год", "Месяц", "Период"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for col, route in enumerate(routes, 5):
        cell = ws.cell(row=1, column=col, value=route)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    cell = ws.cell(row=1, column=total_col, value="Итого")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

    # Строки данных
    data_row = 2
    for period_num, year, month_name, period_str, route_values in monthly_rows:
        ws.cell(row=data_row, column=1, value=period_num).border = thin_border
        ws.cell(row=data_row, column=1).alignment = center_align
        ws.cell(row=data_row, column=2, value=year).border = thin_border
        ws.cell(row=data_row, column=2).alignment = center_align
        ws.cell(row=data_row, column=3, value=month_name).border = thin_border
        ws.cell(row=data_row, column=3).alignment = center_align
        ws.cell(row=data_row, column=4, value=period_str).border = thin_border
        ws.cell(row=data_row, column=4).alignment = center_align

        row_total = 0.0
        for col, route in enumerate(routes, 5):
            val = round(route_values.get(route, 0), 2)
            row_total += val
            cell = ws.cell(row=data_row, column=col, value=val if val > 0 else 0)
            cell.border = thin_border
            cell.alignment = center_align
            cell.number_format = number_format

        cell = ws.cell(row=data_row, column=total_col, value=round(row_total, 2))
        cell.border = thin_border
        cell.alignment = center_align
        cell.number_format = number_format
        cell.font = Font(bold=True)

        data_row += 1

    # Строка ИТОГО
    first_data_row = 2
    last_data_row = data_row - 1
    ws.cell(row=data_row, column=1, value="").border = thin_border
    ws.cell(row=data_row, column=2, value="").border = thin_border
    ws.cell(row=data_row, column=3, value="").border = thin_border
    cell = ws.cell(row=data_row, column=4, value="ИТОГО:")
    cell.border = thin_border
    cell.font = Font(bold=True)
    cell.alignment = center_align

    for col in range(5, total_col + 1):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        cell = ws.cell(row=data_row, column=col, value=formula)
        cell.border = thin_border
        cell.alignment = center_align
        cell.number_format = number_format
        cell.font = Font(bold=True)

    # Ширина столбцов
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    for col in range(5, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12


def _compute_monthly_km_and_rub(
    session: Session,
    contract_id: int,
    routes: list[str],
    route_mapping: dict[str, list[str]],
    start_year: int,
    start_month: int,
    end_year: int,
    end_month: int,
    contract_number: str,
) -> tuple[list[tuple], list[tuple]]:
    """
    Вычисляет помесячные объёмы в км и рублях.

    Returns:
        (km_rows, rub_rows) — каждый список: [(period_num, year, month_name, period_str, {route: value})]
    """
    km_rows = []
    rub_rows = []
    period_num = 1

    current_year = start_year
    current_month = start_month

    while (current_year < end_year) or (current_year == end_year and current_month <= end_month):
        _, last_day = calendar.monthrange(current_year, current_month)
        d_from = date(current_year, current_month, 1)
        d_to = date(current_year, current_month, last_day)

        # Км по маршрутам
        route_km: dict[str, float] = {}
        for norm_route in routes:
            total_km = 0.0
            for orig_route in route_mapping[norm_route]:
                calc = calculate_route_period(session, contract_id, orig_route, d_from, d_to)
                total_km += round(calc.total_km, 2)
            route_km[norm_route] = total_km

        period_str = f"01.{current_month:02d}-{last_day}.{current_month:02d}"
        row_base = (period_num, current_year, MONTH_NAMES[current_month], period_str)
        km_rows.append((*row_base, route_km))

        # Руб по маршрутам
        coefficients = get_coefficients_for_date(d_to, contract_number)
        capacities = get_capacities(contract_number)
        route_rub: dict[str, float] = {}
        for norm_route in routes:
            km = route_km[norm_route]
            normalized = norm_route.rstrip('.')
            coef = coefficients.get(normalized) if coefficients else None
            cap = capacities.get(normalized) if capacities else None
            if coef is not None and cap is not None and km > 0:
                rub = round(float(Decimal(str(km)) * Decimal(str(cap)) * coef), 2)
            else:
                rub = 0.0
            route_rub[norm_route] = rub
        rub_rows.append((*row_base, route_rub))

        period_num += 1
        current_month += 1
        if current_month > 12:
            current_month = 1
            current_year += 1

    return km_rows, rub_rows


def generate_monthly_volumes_rub_excel(
    session: Session,
    contract_id: int,
    contract_number: str = "222",
    start_year: int = 2026,
    start_month: int = 2,
    end_year: int = 2028,
    end_month: int = 7,
    ds_number: Optional[str] = None,
) -> BytesIO:
    """Генерирует Excel файл с помесячными объёмами работ в рублях."""
    routes = get_contract_routes(session, contract_id, normalize=True)
    original_routes = get_contract_routes(session, contract_id, normalize=False)
    route_mapping: dict[str, list[str]] = {}
    for orig in original_routes:
        norm = normalize_route(orig)
        if norm not in route_mapping:
            route_mapping[norm] = []
        route_mapping[norm].append(orig)

    preload_price_data(contract_number)

    _, rub_rows = _compute_monthly_km_and_rub(
        session, contract_id, routes, route_mapping,
        start_year, start_month, end_year, end_month, contract_number
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Объёмы работ, руб."
    _build_monthly_sheet(ws, routes, rub_rows, number_format='#,##0.00')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_monthly_volumes_combined_excel(
    session: Session,
    contract_id: int,
    contract_number: str = "222",
    start_year: int = 2026,
    start_month: int = 2,
    end_year: int = 2028,
    end_month: int = 7,
    ds_number: Optional[str] = None,
) -> BytesIO:
    """Генерирует Excel файл с двумя листами: км и руб."""
    routes = get_contract_routes(session, contract_id, normalize=True)
    original_routes = get_contract_routes(session, contract_id, normalize=False)
    route_mapping: dict[str, list[str]] = {}
    for orig in original_routes:
        norm = normalize_route(orig)
        if norm not in route_mapping:
            route_mapping[norm] = []
        route_mapping[norm].append(orig)

    preload_price_data(contract_number)

    km_rows, rub_rows = _compute_monthly_km_and_rub(
        session, contract_id, routes, route_mapping,
        start_year, start_month, end_year, end_month, contract_number
    )

    wb = Workbook()

    ws_km = wb.active
    ws_km.title = "Объёмы работ, км"
    _build_monthly_sheet(ws_km, routes, km_rows, number_format='#,##0.00')

    ws_rub = wb.create_sheet(title="Объёмы работ, руб.")
    _build_monthly_sheet(ws_rub, routes, rub_rows, number_format='#,##0.00')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_quarterly_volumes_excel(
    session: Session,
    contract_id: int,
    start_year: int = 2022,
    start_quarter: int = 2,
    end_year: int = 2028,
    end_quarter: int = 3,
    ds_number: Optional[str] = None,
) -> BytesIO:
    """
    Генерирует Excel файл с поквартальными объёмами работ (как в образце).

    Формат: кварталы с периодами вида "01.04-30.06"
    """
    # Получаем контракт
    contract = session.query(Contract).filter(Contract.id == contract_id).first()

    # Получаем нормализованный список маршрутов (для отображения в таблице)
    routes = get_contract_routes(session, contract_id, normalize=True)

    # Получаем оригинальные маршруты (для расчёта)
    original_routes = get_contract_routes(session, contract_id, normalize=False)

    # Маппинг: нормализованный -> список оригинальных
    route_mapping: dict[str, list[str]] = {}
    for orig in original_routes:
        norm = normalize_route(orig)
        if norm not in route_mapping:
            route_mapping[norm] = []
        route_mapping[norm].append(orig)

    # Создаём книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Приложение 3 (квартальное)"

    # Стили
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Заголовок
    title = f"Приложение №12"
    if ds_number:
        title += f"\nк дополнительному соглашению №{ds_number}"
    if contract and contract.full_number:
        title += f"\nк государственному контракту {contract.full_number}"

    ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=len(routes) + 6)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # Заголовок таблицы
    header_row = 6
    headers = ["№", "Год", "Квартал", "Период"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Маршруты
    route_row = header_row + 1
    for col in range(1, 5):
        ws.cell(row=route_row, column=col, value="").border = thin_border

    for col, route in enumerate(routes, 5):
        cell = ws.cell(row=route_row, column=col, value=route)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    total_col = len(routes) + 6
    cell = ws.cell(row=route_row, column=total_col, value="Итого")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

    # Кварталы
    quarter_periods = {
        1: ("01.01", "31.03"),
        2: ("01.04", "30.06"),
        3: ("01.07", "30.09"),
        4: ("01.10", "31.12"),
    }

    data_row = route_row + 1
    period_num = 1

    current_year = start_year
    current_quarter = start_quarter

    while (current_year < end_year) or (current_year == end_year and current_quarter <= end_quarter):
        # Период квартала
        start_str, end_str = quarter_periods[current_quarter]
        start_month = (current_quarter - 1) * 3 + 1
        end_month = current_quarter * 3
        _, last_day = calendar.monthrange(current_year, end_month)

        d_from = date(current_year, start_month, 1)
        d_to = date(current_year, end_month, last_day)

        # Рассчитываем объёмы по всем маршрутам с учётом маппинга нормализованных маршрутов
        # Для каждого нормализованного маршрута суммируем данные всех оригинальных маршрутов
        # Округляем каждый маршрут до 2 знаков перед суммированием
        route_calculations = {}
        for norm_route in routes:
            total_km = 0.0
            for orig_route in route_mapping[norm_route]:
                calc = calculate_route_period(session, contract_id, orig_route, d_from, d_to)
                total_km += round(calc.total_km, 2)
            route_calculations[norm_route] = total_km

        # Заполняем строку
        ws.cell(row=data_row, column=1, value=period_num).border = thin_border
        ws.cell(row=data_row, column=1).alignment = center_align

        ws.cell(row=data_row, column=2, value=current_year).border = thin_border
        ws.cell(row=data_row, column=2).alignment = center_align

        ws.cell(row=data_row, column=3, value=current_quarter).border = thin_border
        ws.cell(row=data_row, column=3).alignment = center_align

        period_str = f"{start_str}-{end_str.split('.')[0]}.{end_str.split('.')[1]}"
        ws.cell(row=data_row, column=4, value=period_str).border = thin_border
        ws.cell(row=data_row, column=4).alignment = center_align

        # Объёмы (используем суммированные значения)
        row_total = 0.0
        for col, route in enumerate(routes, 5):
            km = round(route_calculations.get(route, 0), 2)
            row_total += km

            cell = ws.cell(row=data_row, column=col, value=km if km > 0 else 0)
            cell.border = thin_border
            cell.alignment = center_align
            cell.number_format = '#,##0.00'

        cell = ws.cell(row=data_row, column=total_col, value=round(row_total, 2))
        cell.border = thin_border
        cell.alignment = center_align
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)

        # Следующий квартал
        period_num += 1
        data_row += 1
        current_quarter += 1
        if current_quarter > 4:
            current_quarter = 1
            current_year += 1

    # Настройка ширины
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12

    for col in range(5, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
